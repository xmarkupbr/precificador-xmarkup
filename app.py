import os
import xml.etree.ElementTree as ET
import io
import json
from functools import wraps
import click
import re
from flask.cli import with_appcontext
from io import BytesIO
from flask import Flask, jsonify, render_template, request, flash, send_file, redirect, url_for, session
import pandas as pd
from database import init_app, get_db
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from datetime import datetime, timedelta
from flask_mail import Mail, Message
from dotenv import load_dotenv
from itsdangerous import URLSafeTimedSerializer
from collections import Counter
import logging
import requests # ADICIONADO: para scraping, mesmo que não seja usado agora
from bs4 import BeautifulSoup # ADICIONADO: para scraping, mesmo que não seja usado agora
from scraping.simple_scraper import SimplePriceScraper
from scraping.simple_scheduler import SimpleScrapingScheduler
import threading
import time  # ADICIONADO: necessário para o comando start-price-monitor

price_monitor_scheduler = None

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY")

# Configurações de upload de arquivos
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB máximo
app.config['UPLOAD_EXTENSIONS'] = ['.xml', '.xlsx', '.csv']

# Configurar o logger da aplicação
logging.basicConfig(level=logging.INFO) # Define o nível mínimo de log
app.logger.setLevel(logging.INFO) # Garante que o logger da app use INFO ou superior

# --- CONFIGURAÇÃO DO FLASK-MAIL ---
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = ('XMarkup', os.environ.get('MAIL_USERNAME'))

mail = Mail(app)
init_app(app)

# --- CONFIGURAÇÃO DO FLASK-LOGIN ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = "Por favor, faça o login para acessar esta página."
login_manager.login_message_category = "info"

NFE_NAMESPACE = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

# --- TRATAMENTO DE ERROS ---
@app.errorhandler(413)
def too_large(e):
    flash("Arquivo muito grande. O tamanho máximo permitido é 16MB.", "danger")
    return redirect(url_for('precificador'))

def allowed_file(filename):
    if not filename or '.' not in filename:
        return False
    extension = '.' + filename.rsplit('.', 1)[1].lower()
    return extension in app.config['UPLOAD_EXTENSIONS']

# --- DECORATORS E HELPERS ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not getattr(current_user, 'is_admin', False):
            flash("Acesso restrito a administradores.", "danger")
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def get_user_usage(db, user_id, last_reset_date=None):
    start_date = datetime.now() - timedelta(days=30)
    if last_reset_date:
        if isinstance(last_reset_date, str):
            try:
                reset_date = datetime.strptime(last_reset_date, '%Y-%m-%d %H:%M:%S.%f')
            except ValueError:
                reset_date = datetime.strptime(last_reset_date, '%Y-%m-%d %H:%M:%S')
        else:
            reset_date = last_reset_date
        
        if reset_date and reset_date > start_date:
            start_date = reset_date
            
    count = db.execute(
        "SELECT COUNT(id) FROM precificacoes WHERE user_id = ? AND criado_em > ?",
        (user_id, start_date)
    ).fetchone()[0]
    return count

def send_price_alert_email(user_email, user_nome, items_abaixo_custo, prec_id):
    """Envia um e-mail de alerta quando produtos estão com preço abaixo do custo."""
    try:
        msg = Message(
            subject="Alerta de Precificação - Produtos Abaixo do Custo",
            recipients=[user_email]
        )
        msg.html = render_template(
            'emails/price_alert.html', 
            user_nome=user_nome,
            items=items_abaixo_custo,
            prec_id=prec_id
        )
        mail.send(msg)
    except Exception as e:
        print(f"ERRO AO ENVIAR E-MAIL DE ALERTA DE PREÇO: {e}")

# --- FILTROS DE TEMPLATE ---
@app.template_filter("brl")
def format_as_brl(value):
    try:
        float_value = float(value)
        return f"R$ {float_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return value

@app.template_filter("input_num")
def format_for_input_filter(value):
    try:
        return f"{float(value):.2f}"
    except (ValueError, TypeError):
        return value

# --- MODELO DE UTILIZADOR ---
class User(UserMixin):
    def __init__(self, id, email, password, nome_completo, is_admin=False, status_cliente='Safira', 
                 precificacao_limit=5, limit_reset_date=None, default_pricing_method='simple_margin',
                 default_contribution_margin=30.0, default_fixed_costs=0.0, default_monthly_sales_qty=100):
        self.id = id
        self.email = email
        self.password = password
        self.nome_completo = nome_completo
        self.is_admin = is_admin
        self.status_cliente = status_cliente
        self.precificacao_limit = precificacao_limit
        self.limit_reset_date = limit_reset_date
        self.default_pricing_method = default_pricing_method
        self.default_contribution_margin = default_contribution_margin
        self.default_fixed_costs = default_fixed_costs
        self.default_monthly_sales_qty = default_monthly_sales_qty

    def get_reset_token(self):
        s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
        return s.dumps(self.id, salt='password-reset-salt')

    @staticmethod
    def verify_reset_token(token, expires_sec=1800):
        s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
        try:
            user_id = s.loads(token, salt='password-reset-salt', max_age=expires_sec)
        except:
            return None
        db = get_db()
        return db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()

@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    user_data = db.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    if user_data:
        keys = user_data.keys()
        return User(
            id=user_data['id'],
            email=user_data['email'],
            password=user_data['password'],
            nome_completo=user_data['nome_completo'],
            is_admin=user_data['is_admin'] if 'is_admin' in keys else False,
            status_cliente=user_data['status_cliente'] if 'status_cliente' in keys else 'Safira',
            precificacao_limit=user_data['precificacao_limit'] if 'precificacao_limit' in keys else 5,
            limit_reset_date=user_data['limit_reset_date'] if 'limit_reset_date' in keys else None,
            default_pricing_method=user_data['default_pricing_method'] if 'default_pricing_method' in keys else 'simple_margin',
            default_contribution_margin=user_data['default_contribution_margin'] if 'default_contribution_margin' in keys else 30.0,
            default_fixed_costs=user_data['default_fixed_costs'] if 'default_fixed_costs' in keys else 0.0,
            default_monthly_sales_qty=user_data['default_monthly_sales_qty'] if 'default_monthly_sales_qty' in keys else 100
        )
    return None

# --- LÓGICA DE NEGÓCIO ---
def process_nfe_file(xml_file):
    try:
        xml_content = xml_file.read()
        xml_buffer = io.BytesIO(xml_content)
        tree = ET.parse(xml_buffer)
    except ET.ParseError:
        raise ValueError(f"O arquivo '{xml_file.filename}' não é um XML válido ou está corrompido.")
    root = tree.getroot()
    products_data = []
    freight_total = float(root.findtext(".//nfe:total/nfe:ICMSTot/nfe:vFrete", default="0.0", namespaces=NFE_NAMESPACE))
    insurance_total = float(root.findtext(".//nfe:total/nfe:ICMSTot/nfe:vSeg", default="0.0", namespaces=NFE_NAMESPACE))
    other_expenses_total = float(root.findtext(".//nfe:total/nfe:ICMSTot/nfe:vOutro", default="0.0", namespaces=NFE_NAMESPACE))
    discount_total = float(root.findtext(".//nfe:total/nfe:ICMSTot/nfe:vDesc", default="0.0", namespaces=NFE_NAMESPACE))
    nfe_number = root.findtext(".//nfe:ide/nfe:nNF", default="", namespaces=NFE_NAMESPACE).strip()
    nfe_series = root.findtext(".//nfe:ide/nfe:serie", default="", namespaces=NFE_NAMESPACE).strip()
    nf_id = f"{nfe_number}/{nfe_series}" if nfe_number and nfe_series else ""
    
    # NOVAS LINHAS: Extrair valor total da NF e somar valor dos produtos da NF
    total_nfe_value_from_xml = float(root.findtext(".//nfe:total/nfe:ICMSTot/nfe:vNF", default="0.0", namespaces=NFE_NAMESPACE))
    sum_vProd_from_xml_items = 0.0 # Inicializa a soma dos vProd de todos os itens desta NFe

    for det in root.findall(".//nfe:det", NFE_NAMESPACE):
        prod = det.find("nfe:prod", NFE_NAMESPACE)
        imposto = det.find("nfe:imposto", NFE_NAMESPACE)
        if prod is None or imposto is None: continue
        
        item_vprod = float(prod.findtext("nfe:vProd", default="0.0", namespaces=NFE_NAMESPACE))
        sum_vProd_from_xml_items += item_vprod # Soma o vProd de cada item
        
        taxes = sum(float(n.text) for n in [
            imposto.find(".//nfe:ICMS//nfe:vICMS", NFE_NAMESPACE),
            imposto.find(".//nfe:PIS//nfe:vPIS", NFE_NAMESPACE),
            imposto.find(".//nfe:COFINS//nfe:vCOFINS", NFE_NAMESPACE),
        ] if n is not None and n.text is not None)
        products_data.append({
            "Série NF-e": nf_id, "Código": prod.findtext("nfe:cProd", default="N/A", namespaces=NFE_NAMESPACE),
            "Nome": prod.findtext("nfe:xProd", default="N/A", namespaces=NFE_NAMESPACE),
            "Qtd": float(prod.findtext("nfe:qCom", default="0.0", namespaces=NFE_NAMESPACE)),
            "valor_total": item_vprod, # Usar o item_vprod que já extraímos
            "impostos": taxes, "frete_total": freight_total, "seguro_total": insurance_total,
            "outros_total": other_expenses_total, "desc_total": discount_total,
        })
    # RETORNO MODIFICADO:
    return products_data, total_nfe_value_from_xml, sum_vProd_from_xml_items

def process_spreadsheet_file(spreadsheet_file_storage, filename):
    """
    Processa um arquivo de planilha (Excel ou CSV) para extrair dados de produtos.
    Assume que a planilha tem colunas como 'Código', 'Nome', 'Qtd', 'Custo Unitário (R$)'.
    """
    df = None
    file_content = spreadsheet_file_storage.read() # Lê o conteúdo do arquivo
    file_stream = BytesIO(file_content) # Cria um stream de bytes para o pandas

    try:
        if filename.lower().endswith('.xlsx'):
            df = pd.read_excel(file_stream)
        elif filename.lower().endswith('.csv'):
            # Tenta ler com diferentes encodings, comum para CSVs
            encodings_to_try = ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']
            df = None
            last_error = None
            
            for encoding in encodings_to_try:
                try:
                    file_stream.seek(0) # Volta ao início do stream
                    df = pd.read_csv(file_stream, encoding=encoding)
                    app.logger.info(f"Arquivo CSV '{filename}' lido com sucesso usando encoding '{encoding}'")
                    break
                except UnicodeDecodeError as e:
                    last_error = e
                    continue
                except Exception as e:
                    last_error = e
                    break
            
            if df is None:
                raise ValueError(f"Não foi possível ler o arquivo CSV '{filename}' com nenhum dos encodings testados. Último erro: {last_error}")
        else:
            raise ValueError(f"Formato de arquivo não suportado para '{filename}'. Por favor, use .xlsx ou .csv.")
    except Exception as e:
        # Loga o erro completo para depuração
        app.logger.error(f"Erro ao processar planilha '{filename}': {e}", exc_info=True)
        raise ValueError(f"Erro ao ler o arquivo '{filename}': {e}")

    products_data = []
    required_columns = ['Código', 'Nome', 'Qtd', 'Custo Unitário (R$)']
    
    # Verifica se o DataFrame está vazio após a leitura
    if df is None or df.empty:
        raise ValueError(f"Planilha '{filename}' está vazia ou não contém dados válidos.")

    # Remove espaços em branco e caracteres invisíveis dos nomes das colunas
    import re
    df.columns = df.columns.str.strip()  # Remove espaços normais
    # Remove caracteres invisíveis comuns (zero-width space, non-breaking space, etc.)
    df.columns = [re.sub(r'[\u200b\u00a0\ufeff\u2000-\u200f\u2028-\u202f]', '', str(col)) for col in df.columns]
    
    # Validação mais robusta de colunas
    def normalize_column_name(name):
        """Normaliza nome da coluna para comparação"""
        import unicodedata
        # Remove acentos e converte para minúsculas
        normalized = unicodedata.normalize('NFD', str(name).lower())
        normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
        # Remove caracteres especiais exceto parênteses e cifrão
        normalized = re.sub(r'[^\w\s()$]', '', normalized)
        # Remove espaços extras
        normalized = re.sub(r'\s+', ' ', normalized).strip()
        return normalized
    
    # Mapear colunas encontradas para nomes normalizados
    column_mapping = {}
    normalized_found = {}
    
    for col in df.columns:
        normalized = normalize_column_name(col)
        column_mapping[normalized] = col
        normalized_found[normalized] = col
    
    # Definir variações aceitas para cada coluna obrigatória
    required_variations = {
        'Código': ['codigo', 'cod', 'code', 'sku'],
        'Nome': ['nome', 'produto', 'descricao', 'description', 'name'],
        'Qtd': ['qtd', 'quantidade', 'qty', 'quantity', 'quant'],
        'Custo Unitário (R$)': ['custo unitario r$', 'custo unitario', 'custo', 'preco custo', 'valor unitario', 'price']
    }
    
    # Tentar mapear colunas obrigatórias
    mapped_columns = {}
    missing_columns = []
    
    for required_col in required_columns:
        found = False
        
        # Primeiro, tentar match exato
        if required_col in df.columns:
            mapped_columns[required_col] = required_col
            found = True
        else:
            # Tentar match normalizado
            normalized_required = normalize_column_name(required_col)
            if normalized_required in normalized_found:
                mapped_columns[required_col] = normalized_found[normalized_required]
                found = True
            else:
                # Tentar variações conhecidas
                for variation in required_variations.get(required_col, []):
                    if variation in normalized_found:
                        mapped_columns[required_col] = normalized_found[variation]
                        found = True
                        break
        
        if not found:
            missing_columns.append(required_col)
    
    if missing_columns:
        # Gerar sugestões úteis
        suggestions = []
        for missing in missing_columns:
            similar_cols = []
            missing_normalized = normalize_column_name(missing)
            
            for col in df.columns:
                col_normalized = normalize_column_name(col)
                # Verificar se há palavras em comum
                missing_words = set(missing_normalized.split())
                col_words = set(col_normalized.split())
                if missing_words & col_words:  # Interseção não vazia
                    similar_cols.append(col)
            
            if similar_cols:
                suggestions.append(f"'{missing}' → talvez seja: {', '.join(similar_cols)}")
        
        error_msg = f"Planilha '{filename}' não possui todas as colunas necessárias.\n"
        error_msg += f"Faltam: {', '.join(missing_columns)}.\n"
        error_msg += f"Colunas encontradas: {', '.join(df.columns)}.\n"
        if suggestions:
            error_msg += f"Sugestões: {'; '.join(suggestions)}.\n"
        error_msg += "Verifique se os nomes das colunas estão corretos: 'Código', 'Nome', 'Qtd', 'Custo Unitário (R$)'."
        
        app.logger.error(f"Erro de validação de colunas: {error_msg}")
        raise ValueError(error_msg)
    
    # Renomear colunas para os nomes padrão se necessário
    if mapped_columns:
        rename_dict = {v: k for k, v in mapped_columns.items() if v != k}
        if rename_dict:
            df = df.rename(columns=rename_dict)
            app.logger.info(f"Colunas renomeadas na planilha '{filename}': {rename_dict}")

    for index, row in df.iterrows():
        try:
            # Converte os tipos de dados para garantir que são numéricos onde esperado
            qtd = float(row['Qtd'])
            custo_unitario = float(row['Custo Unitário (R$)'])

            # Ignorar linhas com Qtd ou Custo Unitário inválidos/zero
            if qtd <= 0 or custo_unitario <= 0:
                app.logger.warning(f"Linha {index+1} da planilha '{filename}' ignorada: Quantidade ou Custo Unitário inválido/zero.")
                continue

            products_data.append({
                "Série NF-e": f"PLANILHA_{os.path.basename(filename).split('.')[0].upper()}_{index+1}", # Nome mais descritivo
                "Código": str(row['Código']),
                "Nome": str(row['Nome']),
                "Qtd": qtd,
                "valor_total": qtd * custo_unitario, 
                "impostos": 0.0, 
                "frete_total": 0.0, 
                "seguro_total": 0.0,
                "outros_total": 0.0,
                "desc_total": 0.0,
                "Custo Unitário (R$)": custo_unitario
            })
        except ValueError as ve:
            app.logger.error(f"Erro de conversão de dados na linha {index+1} da planilha '{filename}': {ve}. Linha: {row.to_dict()}", exc_info=True)
            # Removido flash() da função - será tratado na rota
            continue
        except KeyError as ke:
            app.logger.error(f"Coluna esperada não encontrada na linha {index+1} da planilha '{filename}': {ke}. Verifique o cabeçalho da planilha.", exc_info=True)
            # Removido flash() da função - será tratado na rota
            continue

    if not products_data:
        raise ValueError(f"Nenhum produto válido foi extraído da planilha '{filename}'. Verifique se as colunas estão corretas e se há dados válidos.")

    return products_data

def calculate_contribution_margin_prices(products_raw, contribution_margin, fixed_costs, monthly_sales_qty, commissions, shipping_costs):
    """
    Calcula preços usando o método de Margem de Contribuição.
    
    Margem de Contribuição = (Preço de Venda - Custos Variáveis) / Preço de Venda
    Preço de Venda = (Custos Variáveis + Rateio de Custos Fixos) / (1 - Margem de Contribuição)
    """
    if not products_raw or monthly_sales_qty <= 0:
        return []
    
    # Calcula o rateio de custos fixos por unidade
    fixed_cost_per_unit = fixed_costs / monthly_sales_qty if monthly_sales_qty > 0 else 0
    
    total_items_value = sum(item["valor_total"] for item in products_raw)
    
    # Garante que as variáveis de despesa são floats
    total_freight_nfe = float(products_raw[0].get('frete_total', 0.0))
    total_insurance_nfe = float(products_raw[0].get('seguro_total', 0.0))
    total_other_nfe = float(products_raw[0].get('outros_total', 0.0))
    total_discount_nfe = float(products_raw[0].get('desc_total', 0.0))

    final_products = []
    for item in products_raw:
        # Lógica para determinar o imposto a ser considerado no custo
        impostos_to_use = item["impostos"]
        is_spreadsheet_source = "PLANILHA_" in item["Série NF-e"]
        
        # Somente aplica a regra se for um XML e tiver os dados de comparação
        if not is_spreadsheet_source and 'total_nfe_value_from_xml' in item and 'sum_vProd_from_xml_items' in item:
            nfe_total_value = item['total_nfe_value_from_xml']
            sum_vProd_nfe = item['sum_vProd_from_xml_items']
            
            # Usar uma pequena tolerância para comparação de floats
            if abs(nfe_total_value - sum_vProd_nfe) < 0.01: # Tolerância de R$ 0.01
                impostos_to_use = 0.0 # Impostos não serão adicionados
        
        # Se for um item de planilha, o "Custo Unitário (R$)" já é o custo final
        if is_spreadsheet_source:
            unit_cost = item["Custo Unitário (R$)"]
        else: # É um item de XML
            proportion = item["valor_total"] / total_items_value if total_items_value > 0 else 0
            item_freight = proportion * total_freight_nfe
            item_insurance = proportion * total_insurance_nfe
            item_other = proportion * total_other_nfe
            item_discount = proportion * total_discount_nfe
            
            # Calcula o custo total do item
            total_cost = (item["valor_total"] + impostos_to_use + item_freight + item_insurance + item_other - item_discount)
            unit_cost = total_cost / item["Qtd"] if item["Qtd"] > 0 else 0
        
        # Calcula o preço usando Margem de Contribuição
        # Preço = (Custo Variável + Custo Fixo Unitário) / (1 - MC%)
        prices = {}
        for channel, commission in commissions.items():
            # Custo variável total = custo unitário + frete do canal
            variable_cost = unit_cost + shipping_costs.get(channel, 0)
            
            # Adiciona o custo fixo unitário
            total_cost_with_fixed = variable_cost + fixed_cost_per_unit
            
            # MC efetiva = MC desejada - comissão do canal
            effective_contribution_margin = contribution_margin - commission
            
            if effective_contribution_margin <= 0 or effective_contribution_margin >= 1:
                # Se a margem efetiva for inválida, usa um preço mínimo
                prices[channel] = total_cost_with_fixed * 2  # Dobra o custo como fallback
            else:
                prices[channel] = total_cost_with_fixed / (1 - effective_contribution_margin)
        
        # Atualiza o custo unitário e os preços de venda no item
        item["Custo Unitário (R$)"] = unit_cost
        item["Preço Venda Site (R$)"] = prices.get("site", 0)
        item["Mercado Livre (R$)"] = prices.get("ml", 0)
        item["Shopee (R$)"] = prices.get("shopee", 0)
        
        final_products.append(item)
    
    return final_products

def calculate_simple_margin_prices(products_raw, margin, commissions, shipping_costs):
    """
    Método original de cálculo usando margem de lucro simples.
    """
    total_items_value = sum(item["valor_total"] for item in products_raw)
    if not products_raw: return []
    
    # Garante que as variáveis de despesa são floats, mesmo que venham como listas por algum erro
    total_freight_nfe = float(products_raw[0].get('frete_total', 0.0))
    total_insurance_nfe = float(products_raw[0].get('seguro_total', 0.0))
    total_other_nfe = float(products_raw[0].get('outros_total', 0.0))
    total_discount_nfe = float(products_raw[0].get('desc_total', 0.0))

    final_products = []
    for item in products_raw:
        # Lógica para determinar o imposto a ser considerado no custo
        impostos_to_use = item["impostos"]
        is_spreadsheet_source = "PLANILHA_" in item["Série NF-e"]
        
        # Somente aplica a regra se for um XML e tiver os dados de comparação
        if not is_spreadsheet_source and 'total_nfe_value_from_xml' in item and 'sum_vProd_from_xml_items' in item:
            nfe_total_value = item['total_nfe_value_from_xml']
            sum_vProd_nfe = item['sum_vProd_from_xml_items']
            
            # Usar uma pequena tolerância para comparação de floats
            # Se o valor total da NFe for aproximadamente igual à soma dos vProd dos itens da NFe,
            # então consideramos que os impostos já estão 'por dentro' e não os somamos novamente.
            if abs(nfe_total_value - sum_vProd_nfe) < 0.01: # Tolerância de R$ 0.01
                impostos_to_use = 0.0 # Impostos não serão adicionados
        
        # Se for um item de planilha, o "Custo Unitário (R$)" já é o custo final
        # e os totais de frete/seguro/outros/desconto da NFe devem ser considerados 0 para este item específico.
        if is_spreadsheet_source:
            unit_cost = item["Custo Unitário (R$)"]
            item_freight = 0.0
            item_insurance = 0.0
            item_other = 0.0
            item_discount = 0.0
        else: # É um item de XML
            proportion = item["valor_total"] / total_items_value if total_items_value > 0 else 0
            item_freight = proportion * total_freight_nfe
            item_insurance = proportion * total_insurance_nfe
            item_other = proportion * total_other_nfe
            item_discount = proportion * total_discount_nfe
            
            # ATENÇÃO: Usar impostos_to_use aqui
            total_cost = (item["valor_total"] + impostos_to_use + item_freight + item_insurance + item_other - item_discount)
            unit_cost = total_cost / item["Qtd"] if item["Qtd"] > 0 else 0
            
        cost_with_margin = unit_cost * (1 + margin)
        
        prices = {channel: (cost_with_margin + shipping_costs.get(channel, 0)) / (1 - commission) if (1 - commission) != 0 else 0
                  for channel, commission in commissions.items()}
        
        # Atualiza o custo unitário e os preços de venda no item
        item["Custo Unitário (R$)"] = unit_cost
        item["Preço Venda Site (R$)"] = prices.get("site", 0)
        item["Mercado Livre (R$)"] = prices.get("ml", 0)
        item["Shopee (R$)"] = prices.get("shopee", 0)
        
        final_products.append(item)
    return final_products

def calculate_product_prices(products_raw, margin, commissions, shipping_costs, pricing_method='simple_margin', 
                           contribution_margin=None, fixed_costs=None, monthly_sales_qty=None):
    """
    Função principal que decide qual método de cálculo usar baseado no pricing_method.
    """
    if pricing_method == 'contribution_margin':
        # Usa o novo método de Margem de Contribuição
        if contribution_margin is None or fixed_costs is None or monthly_sales_qty is None:
            # Se faltar algum parâmetro, usa valores padrão
            contribution_margin = contribution_margin or 0.3  # 30% padrão
            fixed_costs = fixed_costs or 0
            monthly_sales_qty = monthly_sales_qty or 100
        
        return calculate_contribution_margin_prices(
            products_raw, contribution_margin, fixed_costs, 
            monthly_sales_qty, commissions, shipping_costs
        )
    else:
        # Usa o método tradicional de margem simples
        return calculate_simple_margin_prices(products_raw, margin, commissions, shipping_costs)

def generate_summary(products):
    if not products: return {}
    summary = {
        "total_qtd": sum(p.get("Qtd", 0) for p in products),
        "custo_total": sum(p.get("Custo Unitário (R$)", 0) * p.get("Qtd", 0) for p in products),
    }
    channels = ["Preço Venda Site (R$)", "Mercado Livre (R$)", "Shopee (R$)"]
    num_products = len(products)
    for channel in channels:
        total_channel_price = sum(p.get(channel, 0) * p.get("Qtd", 0) for p in products)
        summary[f"{channel}_total"] = total_channel_price
        summary[f"{channel}_media"] = sum(p.get(channel, 0) for p in products) / num_products if num_products > 0 else 0
    return summary

# --- ROTAS PRINCIPAIS E DE AUTENTICAÇÃO ---
@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        # Se alguém fizer POST para a raiz, redirecionar para GET
        # Isso evita o erro 405 e mantém a funcionalidade
        flash("Redirecionamento automático aplicado.", "info")
        return redirect(url_for('home'))
    return render_template('home.html')
    
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        db = get_db()
        user_data = db.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()

        if user_data and user_data['is_deleted']:
            flash('Esta conta foi excluída e não pode ser acedida. Por favor, entre em contacto com o suporte se achar que isto é um erro.', 'danger')
            return redirect(url_for('login'))

        if not user_data or not check_password_hash(user_data['password'], password):
            flash('Email ou senha inválidos. Por favor, tente novamente.', 'danger')
            return redirect(url_for('login'))
        
        keys = user_data.keys()
        user = User(
            id=user_data['id'],
            email=user_data['email'],
            password=user_data['password'],
            nome_completo=user_data['nome_completo'],
            is_admin=user_data['is_admin'] if 'is_admin' in keys else False,
            status_cliente=user_data['status_cliente'] if 'status_cliente' in keys else 'Safira',
            precificacao_limit=user_data['precificacao_limit'] if 'precificacao_limit' in keys else 5,
            limit_reset_date=user_data['limit_reset_date'] if 'limit_reset_date' in keys else None,
            default_pricing_method=user_data['default_pricing_method'] if 'default_pricing_method' in keys else 'simple_margin',
            default_contribution_margin=user_data['default_contribution_margin'] if 'default_contribution_margin' in keys else 30.0,
            default_fixed_costs=user_data['default_fixed_costs'] if 'default_fixed_costs' in keys else 0.0,
            default_monthly_sales_qty=user_data['default_monthly_sales_qty'] if 'default_monthly_sales_qty' in keys else 100
        )
        login_user(user)
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        password_confirm = request.form.get('password_confirm')
        nome_completo = request.form.get('nome_completo')
        empresa = request.form.get('empresa')
        telefone = request.form.get('telefone')
        ramo_atividade = request.form.get('ramo_atividade')
        marketplaces_list = request.form.getlist('marketplaces')
        marketplaces = ','.join(marketplaces_list)
        if password != password_confirm:
            flash('As senhas não coincidem. Por favor, tente novamente.', 'danger')
            return redirect(url_for('register'))
        # Regex para validar: mínimo 8 chars, uma letra, um número, um símbolo
        password_regex = re.compile(r'^(?=.*[A-Za-z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$')
        if not password_regex.match(password):
            flash('A senha não atende aos requisitos de segurança: mínimo de 8 caracteres, contendo letras, números e símbolos (@$!%*?&).', 'danger')
            return redirect(url_for('register'))        
        db = get_db()
        user_exists = db.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
        if user_exists:
            flash('Este email já está cadastrado. Por favor, faça o login.', 'warning')
            return redirect(url_for('login'))
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        sql_query = '''
            INSERT INTO users (email, password, nome_completo, empresa, telefone, ramo_atividade, marketplaces)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        '''
        values = (email, hashed_password, nome_completo, empresa, telefone, ramo_atividade, marketplaces)
        db.execute(sql_query, values)
        db.commit()
        flash('Conta criada com sucesso! Por favor, faça o login.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sessão terminada com sucesso.', 'info')
    return redirect(url_for('home'))

# --- ROTAS DE RECUPERAÇÃO DE SENHA ---
@app.route('/request_reset_password', methods=['GET', 'POST'])
def request_reset_password():
    if request.method == 'POST':
        email = request.form.get('email')
        db = get_db()
        user_data = db.execute('SELECT * FROM users WHERE email = ?', (email,)).fetchone()
        if user_data:
            user = User(id=user_data['id'], email=user_data['email'], password=user_data['password'], nome_completo=user_data['nome_completo'])
            token = user.get_reset_token()
            msg = Message('Redefinição de Senha - XMarkup',
                          recipients=[user.email])
            msg.body = f'''Para redefinir a sua senha, visite o seguinte link:
{url_for('reset_password', token=token, _external=True)}

Se não foi você que fez este pedido, ignore este e-mail. Este link é válido por 30 minutos.
'''
            try:
                mail.send(msg)
                flash('Um e-mail foi enviado com as instruções para redefinir a sua senha.', 'info')
            except Exception as e:
                flash('Ocorreu um erro ao enviar o e-mail. Por favor, tente novamente mais tarde.', 'danger')
                print(f"ERRO DE EMAIL: {e}")
        else:
            flash('Não foi encontrada nenhuma conta com esse e-mail.', 'warning')
        return redirect(url_for('login'))
    return render_template('request_reset.html')

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    user_data = User.verify_reset_token(token)
    if not user_data:
        flash('O link de redefinição é inválido ou expirou.', 'warning')
        return redirect(url_for('request_reset_password'))
    
    if request.method == 'POST':
        password = request.form.get('password')
        password_confirm = request.form.get('password_confirm')
        if password != password_confirm:
            flash('As senhas não coincidem.', 'danger')
            return render_template('reset_password.html')
            
        hashed_password = generate_password_hash(password)
        db = get_db()
        db.execute('UPDATE users SET password = ? WHERE id = ?', (hashed_password, user_data['id']))
        db.commit()
        flash('A sua senha foi atualizada! Pode agora fazer o login.', 'success')
        return redirect(url_for('login'))
        
    return render_template('reset_password.html')

# --- ROTAS DA APLICAÇÃO ---
@app.route('/precificador', methods=["GET", "POST"])
@login_required
def precificador():
    db = get_db()
    contagem_recente = get_user_usage(db, current_user.id, current_user.limit_reset_date)
    limite_atingido = contagem_recente >= current_user.precificacao_limit

    if request.method == "POST":
        if limite_atingido:
            flash(f"Você atingiu o seu limite de {current_user.precificacao_limit} precificações.", "danger")
            return redirect(url_for('dashboard'))

        form_params = request.form.to_dict()
        try:
            uploaded_files = request.files.getlist("xmlfiles") 
            if not uploaded_files:
                flash("Por favor, selecione ao menos um arquivo XML ou planilha.", "danger")
                return redirect(url_for('precificador'))

            # Filtrar arquivos vazios e validar extensões
            valid_files = []
            for f in uploaded_files:
                if f.filename and f.filename.strip() != '':
                    if allowed_file(f.filename):
                        valid_files.append(f)
                    else:
                        flash(f"Arquivo '{f.filename}' tem extensão não permitida. Use apenas .xml, .xlsx ou .csv.", "warning")
            
            if not valid_files:
                flash("Por favor, selecione ao menos um arquivo válido (.xml, .xlsx ou .csv).", "danger")
                return redirect(url_for('precificador'))

            plataformas = ['site', 'ml', 'shopee']
            
            # Obtém o método de precificação escolhido
            pricing_method = form_params.get("pricing_method", "simple_margin")
            
            # Parâmetros comuns
            comissoes = {p: float(form_params.get(f"comissao_{p}", "0").replace(',', '.')) / 100 for p in plataformas}
            fretes = {p: float(form_params.get(f"frete_{p}", "0").replace(',', '.')) for p in plataformas}
            
            # Parâmetros específicos por método
            if pricing_method == "contribution_margin":
                # Margem de Contribuição
                contribution_margin = float(form_params.get("contribution_margin", "30").replace(',', '.')) / 100
                fixed_costs = float(form_params.get("fixed_costs", "0").replace(',', '.'))
                monthly_sales_qty = int(form_params.get("monthly_sales_qty", "100"))
                margin = None  # Não usado neste método
            else:
                # Margem Simples
                margem = float(form_params.get("margem", "0").replace(',', '.')) / 100
                contribution_margin = None
                fixed_costs = None
                monthly_sales_qty = None
                margin = margem
            
            raw_products = []
            processed_files = 0
            
            for uploaded_file in valid_files:
                filename = uploaded_file.filename
                filename_lower = filename.lower()
                
                # Reset do ponteiro do arquivo para garantir leitura do início
                uploaded_file.seek(0) 

                if filename_lower.endswith('.xml'):
                    try:
                        # RETORNO MODIFICADO: Recebe 3 valores da função process_nfe_file
                        products, total_nfe_value, sum_vProd_from_xml_items = process_nfe_file(uploaded_file)
                        # Adiciona os valores da NFe a cada produto para que a lógica de cálculo possa acessá-los
                        for p in products:
                            p['total_nfe_value_from_xml'] = total_nfe_value
                            p['sum_vProd_from_xml_items'] = sum_vProd_from_xml_items
                        raw_products.extend(products)
                        processed_files += 1
                        app.logger.info(f"XML '{filename}' processado com sucesso. {len(products)} produtos encontrados.")
                    except ValueError as e:
                        flash(f"Erro ao processar XML '{filename}': {e}", "danger")
                        app.logger.error(f"Erro ao processar XML '{filename}': {e}", exc_info=True)
                        continue  # Continua com outros arquivos em vez de retornar
                elif filename_lower.endswith('.xlsx') or filename_lower.endswith('.csv'):
                    try:
                        products = process_spreadsheet_file(uploaded_file, filename)
                        # Para planilhas, define valores padrão ou irrelevantes para as novas chaves
                        for p in products:
                            p['total_nfe_value_from_xml'] = 0.0 
                            p['sum_vProd_from_xml_items'] = 0.0 
                        raw_products.extend(products)
                        processed_files += 1
                        app.logger.info(f"Planilha '{filename}' processada com sucesso. {len(products)} produtos encontrados.")
                    except ValueError as e:
                        flash(f"Erro ao processar planilha '{filename}': {e}", "danger")
                        app.logger.error(f"Erro ao processar planilha '{filename}': {e}", exc_info=True)
                        continue  # Continua com outros arquivos em vez de retornar
                else:
                    flash(f"Formato de arquivo não suportado: {filename}. Por favor, use .xml, .xlsx ou .csv.", "warning")
                    continue

            if processed_files == 0:
                flash("Nenhum arquivo foi processado com sucesso. Verifique os formatos e conteúdo dos arquivos.", "danger")
                return redirect(url_for('precificador'))

            if not raw_products:
                flash("Nenhum produto válido foi encontrado nos arquivos processados. Verifique se os arquivos não estão vazios ou corrompidos, e se as colunas estão corretas.", "danger")
                return redirect(url_for('precificador'))

            # Chama a função de cálculo com os parâmetros corretos
            final_products = calculate_product_prices(
                raw_products, 
                margin=margin,
                commissions=comissoes, 
                shipping_costs=fretes,
                pricing_method=pricing_method,
                contribution_margin=contribution_margin,
                fixed_costs=fixed_costs,
                monthly_sales_qty=monthly_sales_qty
            )
            
            dados_json = json.dumps(final_products)
            parametros_json = json.dumps(form_params)
            
            cursor = db.cursor()
            cursor.execute("INSERT INTO precificacoes (user_id, dados_json, parametros_json) VALUES (?, ?, ?)",
                           (current_user.id, dados_json, parametros_json))
            new_id = cursor.lastrowid
            db.commit()
            
            flash(f"Precificação criada com sucesso! {len(final_products)} produtos processados de {processed_files} arquivo(s).", "success")
            return redirect(url_for('ver_precificacao', prec_id=new_id))
        except ValueError as e:
            flash(f"Erro nos dados enviados: {e}", "danger")
            # Este catch serve para erros genéricos de ValueError que não foram tratados nas funções de processamento de arquivo
            app.logger.error(f"Erro de validação na rota precificador: {e}", exc_info=True)
            return redirect(url_for('precificador'))
        except Exception as e:
            flash(f"Ocorreu um erro inesperado: {e}", "danger")
            app.logger.error(f"Erro inesperado no precificador: {e}", exc_info=True)
            return redirect(url_for('precificador'))
            
    user_settings = db.execute('SELECT * FROM users WHERE id = ?', (current_user.id,)).fetchone()
    parametros = {
        'margem': user_settings['default_margem'],
        'comissao_site': user_settings['default_comissao_site'],
        'frete_site': user_settings['default_frete_site'],
        'comissao_ml': user_settings['default_comissao_ml'],
        'frete_ml': user_settings['default_frete_ml'],
        'comissao_shopee': user_settings['default_comissao_shopee'],
        'frete_shopee': user_settings['default_frete_shopee'],
        'contribution_margin': user_settings['default_contribution_margin'],
    }
    return render_template(
        "precificador.html", 
        parametros=parametros,
        limite_atingido=limite_atingido, 
        contagem_recente=contagem_recente,
        limite_total=current_user.precificacao_limit
    )
    
# --- FUNÇÕES AUXILIARES PARA TEMPLATES ---
def create_excel_template():
    """Cria o arquivo modelo Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    
    # Cabeçalhos
    headers = ['Código', 'Nome', 'Qtd', 'Custo Unitário (R$)']
    ws.append(headers)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7e41c4", end_color="7e41c4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Aplicar estilos aos cabeçalhos
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Adicionar dados de exemplo
    examples = [
        ['001', 'Produto Exemplo 1', 10, 25.50],
        ['002', 'Produto Exemplo 2', 5, 40.00],
        ['003', 'Produto Exemplo 3', 15, 15.75],
    ]
    
    for row_data in examples:
        ws.append(row_data)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def create_csv_template():
    """Cria o DataFrame modelo para CSV"""
    data = {
        'Código': ['001', '002', '003'],
        'Nome': ['Produto Exemplo 1', 'Produto Exemplo 2', 'Produto Exemplo 3'],
        'Qtd': [10, 5, 15],
        'Custo Unitário (R$)': [25.5, 40.0, 15.75]
    }
    return pd.DataFrame(data)

# --- ROTAS DE DOWNLOAD DE MODELOS ---
@app.route('/download/modelo-excel')
@login_required
def download_modelo_excel():
    """Rota para download do arquivo modelo Excel"""
    try:
        template_path = os.path.join(app.root_path, 'static', 'templates', 'modelo_produtos.xlsx')
        
        # Verificar se o arquivo existe
        if not os.path.exists(template_path):
            app.logger.info(f"Arquivo modelo Excel não encontrado em {template_path}, criando dinamicamente...")
            
            # Criar diretório se não existir
            os.makedirs(os.path.dirname(template_path), exist_ok=True)
            
            # Criar o arquivo
            wb = create_excel_template()
            wb.save(template_path)
            app.logger.info(f"Arquivo modelo Excel criado com sucesso em {template_path}")
        
        # Fazer download do arquivo
        return send_file(
            template_path,
            as_attachment=True,
            download_name='XMarkup_Modelo_Produtos.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        app.logger.error(f"Erro inesperado ao baixar modelo Excel: {e}", exc_info=True)
        flash("Erro inesperado ao baixar arquivo modelo Excel. Tente novamente.", "danger")
        return redirect(url_for('precificador'))

@app.route('/download/modelo-csv')
@login_required
def download_modelo_csv():
    """Rota para download do arquivo modelo CSV"""
    try:
        template_path = os.path.join(app.root_path, 'static', 'templates', 'modelo_produtos.csv')
        
        # Verificar se o arquivo existe
        if not os.path.exists(template_path):
            app.logger.info(f"Arquivo modelo CSV não encontrado em {template_path}, criando dinamicamente...")
            
            # Criar diretório se não existir
            os.makedirs(os.path.dirname(template_path), exist_ok=True)
            
            # Criar o arquivo
            df = create_csv_template()
            df.to_csv(template_path, index=False, encoding='utf-8')
            app.logger.info(f"Arquivo modelo CSV criado com sucesso em {template_path}")
        
        # Fazer download do arquivo
        return send_file(
            template_path,
            as_attachment=True,
            download_name='XMarkup_Modelo_Produtos.csv',
            mimetype='text/csv'
        )
        
    except Exception as e:
        app.logger.error(f"Erro inesperado ao baixar modelo CSV: {e}", exc_info=True)
        flash("Erro inesperado ao baixar arquivo modelo CSV. Tente novamente.", "danger")
        return redirect(url_for('precificador'))

# --- ROTAS PÚBLICAS ALTERNATIVAS PARA DOWNLOAD DE MODELOS ---
@app.route('/public/modelo-excel')
def public_download_modelo_excel():
    """Rota pública para download do arquivo modelo Excel (sem login)"""
    try:
        template_path = os.path.join(app.root_path, 'static', 'templates', 'modelo_produtos.xlsx')
        
        # Verificar se o arquivo existe
        if not os.path.exists(template_path):
            app.logger.info(f"Arquivo modelo Excel não encontrado em {template_path}, criando dinamicamente...")
            
            # Criar diretório se não existir
            os.makedirs(os.path.dirname(template_path), exist_ok=True)
            
            # Criar o arquivo
            wb = create_excel_template()
            wb.save(template_path)
            app.logger.info(f"Arquivo modelo Excel criado com sucesso em {template_path}")
        
        # Fazer download do arquivo
        return send_file(
            template_path,
            as_attachment=True,
            download_name='XMarkup_Modelo_Produtos.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        app.logger.error(f"Erro inesperado ao baixar modelo Excel: {e}", exc_info=True)
        return f"Erro inesperado ao baixar arquivo modelo Excel: {e}", 500

@app.route('/public/modelo-csv')
def public_download_modelo_csv():
    """Rota pública para download do arquivo modelo CSV (sem login)"""
    try:
        template_path = os.path.join(app.root_path, 'static', 'templates', 'modelo_produtos.csv')
        
        # Verificar se o arquivo existe
        if not os.path.exists(template_path):
            app.logger.info(f"Arquivo modelo CSV não encontrado em {template_path}, criando dinamicamente...")
            
            # Criar diretório se não existir
            os.makedirs(os.path.dirname(template_path), exist_ok=True)
            
            # Criar o arquivo
            df = create_csv_template()
            df.to_csv(template_path, index=False, encoding='utf-8')
            app.logger.info(f"Arquivo modelo CSV criado com sucesso em {template_path}")
        
        # Fazer download do arquivo
        return send_file(
            template_path,
            as_attachment=True,
            download_name='XMarkup_Modelo_Produtos.csv',
            mimetype='text/csv'
        )
        
    except Exception as e:
        app.logger.error(f"Erro inesperado ao baixar modelo CSV: {e}", exc_info=True)
        return f"Erro inesperado ao baixar arquivo modelo CSV: {e}", 500

@app.route('/dashboard')
@login_required
def dashboard():
    db = get_db()
    
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')
    
    query_params = [current_user.id]
    date_filter_query = ""
    
    if start_date_str and end_date_str:
        end_date_dt = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1, seconds=-1)
        date_filter_query = "AND p.criado_em BETWEEN ? AND ?"
        query_params.extend([start_date_str, end_date_dt])

    base_query = f"""
        SELECT p.id, p.criado_em, p.dados_json 
        FROM precificacoes p 
        WHERE p.user_id = ? {date_filter_query}
        ORDER BY p.criado_em DESC
    """
    all_precificacoes_raw = db.execute(base_query, tuple(query_params)).fetchall()
    
    search_nfe = request.args.get('search_nfe', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 10 
    
    filtered_precificacoes = []
    if search_nfe:
        for row in all_precificacoes_raw:
            dados = json.loads(row['dados_json'])
            if dados and any(item.get("Série NF-e", "") == search_nfe for item in dados):
                filtered_precificacoes.append(row)
    else:
        filtered_precificacoes = all_precificacoes_raw
        
    total_items = len(filtered_precificacoes)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_precificacoes = filtered_precificacoes[start:end]
    total_pages = (total_items + per_page - 1) // per_page
    
    kpis = {'custo_total': 0, 'receita_total': 0, 'lucro_total': 0, 'total_itens': 0}
    receita_por_canal = {'Meu Site': 0, 'Mercado Livre': 0, 'Shopee': 0}
    
    for row in all_precificacoes_raw:
        dados = json.loads(row['dados_json'])
        for produto in dados:
            custo = produto.get('Custo Unitário (R$)', 0) * produto.get('Qtd', 0)
            kpis['custo_total'] += custo
            kpis['total_itens'] += produto.get('Qtd', 0)
            receita_site = produto.get('Preço Venda Site (R$)', 0) * produto.get('Qtd', 0)
            receita_ml = produto.get('Mercado Livre (R$)', 0) * produto.get('Qtd', 0)
            receita_shopee = produto.get('Shopee (R$)', 0) * produto.get('Qtd', 0)
            receita_por_canal['Meu Site'] += receita_site
            receita_por_canal['Mercado Livre'] += receita_ml
            receita_por_canal['Shopee'] += receita_shopee
            
    kpis['receita_total'] = sum(receita_por_canal.values())
    kpis['lucro_total'] = kpis['receita_total'] - kpis['custo_total']
    
    chart_data = {
        'labels': list(receita_por_canal.keys()),
        'data': list(receita_por_canal.values())
    }
    
    precificacoes_list = []
    for row in paginated_precificacoes:
        dados = json.loads(row['dados_json'])
        resumo = generate_summary(dados)
        nfe_num = dados[0].get("Série NF-e", "N/A") if dados else "N/A"
        precificacoes_list.append({
            'id': row['id'],
            'criado_em': datetime.strptime(row['criado_em'], '%Y-%m-%d %H:%M:%S'),
            'nfe': nfe_num,
            'num_produtos': len(dados),
            'custo_total': resumo.get('custo_total', 0)
        })
        
    return render_template(
        'dashboard.html', 
        precificacoes=precificacoes_list, 
        kpis=kpis, 
        chart_data=chart_data,
        page=page,
        total_pages=total_pages,
        search_nfe=search_nfe,
        start_date=start_date_str,
        end_date=end_date_str
    )

@app.route('/precificacao/<int:prec_id>')
@login_required
def ver_precificacao(prec_id):
    db = get_db()
    query = "SELECT dados_json, parametros_json FROM precificacoes WHERE id = ?"
    params = (prec_id,)
    if not getattr(current_user, 'is_admin', False):
        query += " AND user_id = ?"
        params = (prec_id, current_user.id)
    prec_data = db.execute(query, params).fetchone()
    if prec_data is None:
        flash("Precificação não encontrada ou acesso não permitido.", "danger")
        return redirect(url_for('dashboard'))
    page = request.args.get('page', 1, type=int)
    per_page = 20
    produtos_todos = json.loads(prec_data['dados_json'])
    parametros = json.loads(prec_data['parametros_json']) if prec_data['parametros_json'] else {}
    resumo = generate_summary(produtos_todos)
    session['precificacao_id'] = prec_id
    total_items = len(produtos_todos)
    start = (page - 1) * per_page
    end = start + per_page
    produtos_paginados = produtos_todos[start:end]
    total_pages = (total_items + per_page - 1) // per_page
    return render_template(
        'precificador.html', 
        produtos=produtos_paginados,
        resumo=resumo, 
        parametros=parametros,
        page=page,
        total_pages=total_pages,
        prec_id=prec_id
    )

@app.route('/api/precificacao/<int:prec_id>/salvar', methods=['POST'])
@login_required
def salvar_precificacao_ajax(prec_id):
    data = request.get_json()
    if not data:
        return jsonify({'status': 'error', 'message': 'Nenhum dado recebido.'}), 400

    produtos_atualizados = data.get('produtos')
    parametros_atualizados = data.get('parametros')

    if not produtos_atualizados or not parametros_atualizados:
        return jsonify({'status': 'error', 'message': 'Dados da requisição estão incompletos.'}), 400

    db = get_db()

    prec_record = db.execute(
        "SELECT id FROM precificacoes WHERE id = ? AND user_id = ?",
        (prec_id, current_user.id)
    ).fetchone()

    if not prec_record:
        return jsonify({'status': 'error', 'message': 'Precificação não encontrada ou acesso não permitido.'}), 403

    # Verificação do método de precificação ao recalcular
    pricing_method = parametros_atualizados.get('pricing_method', 'simple_margin')
    
    items_abaixo_custo = []
    for produto in produtos_atualizados:
        custo_unitario = float(produto.get('Custo Unitário (R$)', 0))
        precos = {
            "Meu Site": float(produto.get('Preço Venda Site (R$)', 0)),
            "Mercado Livre": float(produto.get('Mercado Livre (R$)', 0)),
            "Shopee": float(produto.get('Shopee (R$)', 0))
        }
        for canal, preco_venda in precos.items():
            if preco_venda > 0 and preco_venda < custo_unitario:
                items_abaixo_custo.append({
                    "codigo": produto.get("Código"),
                    "nome": produto.get("Nome"),
                    "custo": custo_unitario,
                    "preco": preco_venda,
                    "canal": canal
                })
    
    if items_abaixo_custo:
        send_price_alert_email(current_user.email, current_user.nome_completo, items_abaixo_custo, prec_id)

    try:
        dados_json_str = json.dumps(produtos_atualizados)
        parametros_json_str = json.dumps(parametros_atualizados)

        db.execute(
            "UPDATE precificacoes SET dados_json = ?, parametros_json = ? WHERE id = ?",
            (dados_json_str, parametros_json_str, prec_id)
        )
        db.commit()

        return jsonify({'status': 'success', 'message': 'Alterações guardadas com sucesso!'})

    except Exception as e:
        print(f"ERRO AO SALVAR PRECIFICACAO (ID: {prec_id}): {e}")
        return jsonify({'status': 'error', 'message': 'Ocorreu um erro interno ao tentar guardar as alterações.'}), 500

@app.route('/perfil')
@login_required
def perfil():
    db = get_db()
    user_data = db.execute(
        'SELECT * FROM users WHERE id = ?', (current_user.id,)
    ).fetchone()
    # Correção: O user_data['limit_reset_date'] pode ser None ou uma string.
    # A função get_user_usage já lida com o tipo.
    uso_recente = get_user_usage(db, current_user.id, user_data['limit_reset_date'])
    return render_template(
        'perfil.html', 
        user=user_data, 
        uso_recente=uso_recente
    )

@app.route('/perfil/editar', methods=['GET', 'POST'])
@login_required
def editar_perfil():
    db = get_db()
    if request.method == 'POST':
        nome_completo = request.form.get('nome_completo')
        empresa = request.form.get('empresa')
        telefone = request.form.get('telefone')
        ramo_atividade = request.form.get('ramo_atividade')
        marketplaces_list = request.form.getlist('marketplaces')
        marketplaces = ','.join(marketplaces_list)
        db.execute(
            '''UPDATE users SET 
               nome_completo = ?, empresa = ?, telefone = ?, ramo_atividade = ?, marketplaces = ?
               WHERE id = ?''',
            (nome_completo, empresa, telefone, ramo_atividade, marketplaces, current_user.id)
        )
        db.commit()
        flash('Perfil atualizado com sucesso!', 'success')
        return redirect(url_for('perfil'))
    user_data = db.execute(
        'SELECT * FROM users WHERE id = ?', (current_user.id,)
    ).fetchone()
    return render_template('editar_perfil.html', user=user_data)

@app.route('/perfil/excluir', methods=['POST'])
@login_required
def excluir_conta():
    reason = request.form.get('delete_reason')
    
    if not reason or len(reason) < 100:
        flash('É necessário fornecer um motivo com pelo menos 100 caracteres para excluir a conta.', 'danger')
        return redirect(url_for('perfil'))

    db = get_db()
    user_id = current_user.id
    user_email = current_user.email
    user_nome = current_user.nome_completo
    
    db.execute(
        'UPDATE users SET is_deleted = 1, deleted_at = ?, delete_reason = ? WHERE id = ?',
        (datetime.now(), reason, user_id)
    )
    
    db.commit()

    try:
        msg = Message(
            subject=f"Notificação: Conta Excluída - {user_email}",
            recipients=[os.environ.get('MAIL_USERNAME')]
        )
        msg.body = f"""
O utilizador abaixo excluiu a sua conta do XMarkup.

- Nome: {user_nome}
- Email: {user_email}
- Data da Exclusão: {datetime.now().strftime('%d/%m/%Y %H:%M')}

Motivo fornecido:
-----------------
{reason}
-----------------

Pode ver mais detalhes no painel de administração.
"""
        mail.send(msg)
    except Exception as e:
        print(f"ERRO AO ENVIAR E-MAIL DE NOTIFICAÇÃO DE EXCLUSÃO: {e}")

    logout_user()
    
    flash('A sua conta foi excluída com sucesso. Lamentamos vê-lo partir!', 'success')
    return redirect(url_for('home'))
    
@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    db = get_db()
    if request.method == 'POST':
        try:
            params_to_update = {
                'default_margem': float(request.form.get('default_margem', 0)),
                'default_comissao_site': float(request.form.get('default_comissao_site', 0)),
                'default_frete_site': float(request.form.get('default_frete_site', 0)),
                'default_comissao_ml': float(request.form.get('default_comissao_ml', 0)),
                'default_frete_ml': float(request.form.get('default_frete_ml', 0)),
                'default_comissao_shopee': float(request.form.get('default_comissao_shopee', 0)),
                'default_frete_shopee': float(request.form.get('default_frete_shopee', 0)),
                # Novos parâmetros para Margem de Contribuição
                'default_pricing_method': request.form.get('default_pricing_method', 'simple_margin'),
                'default_contribution_margin': float(request.form.get('default_contribution_margin', 30)),
                'default_fixed_costs': float(request.form.get('default_fixed_costs', 0)),
                'default_monthly_sales_qty': int(request.form.get('default_monthly_sales_qty', 100)),
            }
            query = "UPDATE users SET " + ", ".join([f"{key} = ?" for key in params_to_update.keys()]) + " WHERE id = ?"
            values = list(params_to_update.values()) + [current_user.id]
            db.execute(query, tuple(values))
            db.commit()
            flash('Configurações salvas com sucesso!', 'success')
        except ValueError:
            flash('Erro: Por favor, insira apenas números válidos.', 'danger')
        return redirect(url_for('settings'))
    user_settings = db.execute('SELECT * FROM users WHERE id = ?', (current_user.id,)).fetchone()
    return render_template('settings.html', user_settings=user_settings)

# --- ROTAS DE GESTÃO DE CONCORRENTES (Fase 1: Configurações) ---
@app.route('/settings/competitors', methods=['GET', 'POST'])
@login_required
def settings_competitors():
    db = get_db()
    if request.method == 'POST':
        competitor_name = request.form.get('name').strip()
        website_url = request.form.get('website_url').strip()
        ml_url = request.form.get('ml_url').strip()
        shopee_url = request.form.get('shopee_url').strip()
        amazon_url = request.form.get('amazon_url').strip()
        
        if not competitor_name:
            flash('O nome do concorrente é obrigatório.', 'danger')
        else:
            try:
                cursor = db.cursor()
                # Verifica se o concorrente já existe para este utilizador
                existing_competitor = cursor.execute(
                    'SELECT id FROM competitors WHERE user_id = ? AND name = ?',
                    (current_user.id, competitor_name)
                ).fetchone()

                if existing_competitor:
                    flash(f"Um concorrente com o nome '{competitor_name}' já existe.", "warning")
                else:
                    cursor.execute(
                        '''INSERT INTO competitors (user_id, name, website_url, ml_url, shopee_url, amazon_url)
                           VALUES (?, ?, ?, ?, ?, ?)''',
                        (current_user.id, competitor_name, website_url if website_url else None,
                         ml_url if ml_url else None, shopee_url if shopee_url else None,
                         amazon_url if amazon_url else None)
                    )
                    db.commit()
                    flash(f"Concorrente '{competitor_name}' adicionado com sucesso!", "success")
            except Exception as e:
                flash(f"Erro ao adicionar concorrente: {e}", "danger")
                db.rollback()
        return redirect(url_for('settings_competitors'))

    competitors = db.execute(
        'SELECT * FROM competitors WHERE user_id = ? ORDER BY name',
        (current_user.id,)
    ).fetchall()
    
    return render_template('settings_competitors.html', competitors=competitors)

@app.route('/settings/competitors/edit/<int:competitor_id>', methods=['POST'])
@login_required
def edit_competitor(competitor_id):
    db = get_db()
    competitor = db.execute(
        'SELECT * FROM competitors WHERE id = ? AND user_id = ?',
        (competitor_id, current_user.id)
    ).fetchone()

    if not competitor:
        print(f"DEBUG: Concorrente {competitor_id} não encontrado ou não autorizado para o utilizador {current_user.id}")
        flash('Concorrente não encontrado ou acesso não permitido.', 'danger')
        return redirect(url_for('settings_competitors'))
    
    new_name = request.form.get('name').strip()
    new_website_url = request.form.get('website_url').strip()
    new_ml_url = request.form.get('ml_url').strip()
    new_shopee_url = request.form.get('shopee_url').strip()
    new_amazon_url = request.form.get('amazon_url').strip()

    print(f"DEBUG: Tentando editar concorrente {competitor_id}")
    print(f"DEBUG: Novos dados: Nome={new_name}, Site={new_website_url}, ML={new_ml_url}, Shopee={new_shopee_url}, Amazon={new_amazon_url}")

    if not new_name:
        print("DEBUG: Nome do concorrente vazio.")
        flash('O nome do concorrente não pode ser vazio.', 'danger')
    else:
        try:
            cursor = db.cursor()
            # Verifica se o novo nome já existe para outro concorrente do mesmo utilizador
            existing_with_new_name = cursor.execute(
                'SELECT id FROM competitors WHERE user_id = ? AND name = ? AND id != ?',
                (current_user.id, new_name, competitor_id)
            ).fetchone()
            
            if existing_with_new_name:
                print(f"DEBUG: Já existe concorrente com o nome '{new_name}'.")
                flash(f"Já existe um concorrente com o nome '{new_name}'.", "warning")
            else:
                cursor.execute(
                    '''UPDATE competitors SET name = ?, website_url = ?, ml_url = ?, shopee_url = ?, amazon_url = ?
                       WHERE id = ? AND user_id = ?''',
                    (new_name, new_website_url if new_website_url else None,
                     new_ml_url if new_ml_url else None, new_shopee_url if new_shopee_url else None,
                     new_amazon_url if new_amazon_url else None,
                     competitor_id, current_user.id)
                )
                db.commit()
                print(f"DEBUG: Concorrente {new_name} atualizado com sucesso no DB.")
                flash(f"Concorrente '{new_name}' atualizado com sucesso!", "success")
        except Exception as e:
            print(f"DEBUG: Erro na atualização do DB: {e}")
            flash(f"Erro ao atualizar concorrente: {e}", "danger")
            db.rollback()
    
    # Adicione este print para ver se o redirecionamento é alcançado
    print("DEBUG: Fim da rota de edição, redirecionando...")
    return redirect(url_for('settings_competitors'))


@app.route('/settings/competitors/delete/<int:competitor_id>', methods=['POST'])
@login_required
def delete_competitor(competitor_id):
    db = get_db()
    try:
        # Garante que o utilizador só pode apagar os seus próprios concorrentes
        # ON DELETE CASCADE na tabela competitor_products (se adicionado) cuidaria dos produtos
        # Mas vamos fazer explicitamente para garantir e dar feedback
        cursor = db.cursor()
        
        # Pega o nome do concorrente para a flash message
        competitor_name_row = cursor.execute(
            'SELECT name FROM competitors WHERE id = ? AND user_id = ?',
            (competitor_id, current_user.id)
        ).fetchone()

        if not competitor_name_row:
            flash('Concorrente não encontrado ou acesso não permitido.', 'danger')
            return redirect(url_for('settings_competitors'))
        
        competitor_name = competitor_name_row['name']

        # Opcional: deletar produtos vinculados manualmente se ON DELETE CASCADE não for configurado ou se quiser contar
        # num_products_deleted = cursor.execute('DELETE FROM competitor_products WHERE competitor_profile_id = ?', (competitor_id,)).rowcount
        
        cursor.execute('DELETE FROM competitors WHERE id = ? AND user_id = ?', (competitor_id, current_user.id))
        db.commit()
        flash(f"Concorrente '{competitor_name}' e seus produtos associados foram excluídos com sucesso!", "success")
    except Exception as e:
        flash(f"Erro ao excluir concorrente: {e}", "danger")
        db.rollback()
    return redirect(url_for('settings_competitors'))


@app.route('/sobre')
def sobre():
    return render_template('sobre.html')

@app.route('/feedback', methods=["GET", "POST"])
def feedback():
    if request.method == "POST":
        email = request.form.get('email')
        message = request.form.get('feedback_message')
        msg = Message(
            subject=f"Novo Feedback de {email or 'Anônimo'}",
            recipients=[os.environ.get('MAIL_USERNAME')]
        )
        msg.body = f"""
        Um novo feedback foi enviado através do XMarkup.

        De: {email or 'Não informado'}
        Mensagem:
        -----------------
        {message}
        -----------------
        """
        try:
            mail.send(msg)
            flash("Obrigado pelo seu feedback! A sua mensagem foi enviada.", "success")
        except Exception as e:
            print(f"ERRO AO ENVIAR E-MAIL DE FEEDBACK: {e}")
            flash("Ocorreu um erro ao tentar enviar a sua mensagem. Por favor, tente novamente mais tarde.", "danger")
        return redirect(url_for('feedback')) # LINHA ALTERADA AQUI
    else:
        return render_template('sobre.html')

# --- ROTAS DE ADMIN ---
@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    db = get_db()
    all_users_raw = db.execute("SELECT * FROM users ORDER BY nome_completo").fetchall()
    
    all_users = []
    for user_row in all_users_raw:
        user = dict(user_row)
        keys = user_row.keys()
        reset_date = user_row['limit_reset_date'] if 'limit_reset_date' in keys else None
        user['uso_recente'] = get_user_usage(db, user['id'], reset_date)
        all_users.append(user)

    all_precificacoes = db.execute("""
        SELECT p.id, p.criado_em, p.dados_json, u.nome_completo as user_nome
        FROM precificacoes p JOIN users u ON p.user_id = u.id
        ORDER BY p.criado_em DESC
        LIMIT 10
    """).fetchall()
    
    precificacoes_list = []
    for row in all_precificacoes:
        dados = json.loads(row['dados_json'])
        nfe_num = dados[0].get("Série NF-e", "N/A") if dados else "N/A"
        precificacoes_list.append({
            'id': row['id'],
            'user_nome': row['user_nome'],
            'criado_em': datetime.strptime(row['criado_em'], '%Y-%m-%d %H:%M:%S'),
            'num_produtos': len(dados),
            'nfe': nfe_num,
            'custo_total': generate_summary(dados).get('custo_total', 0)
        })

    kpis = {
        'total_users': len(all_users),
        'total_precificacoes': db.execute("SELECT COUNT(id) FROM precificacoes").fetchone()[0]
    }
    return render_template(
        'admin_dashboard.html', 
        users=all_users, 
        precificacoes=precificacoes_list,
        kpis=kpis
    )

@app.route('/admin/cliente/<int:user_id>')
@login_required
@admin_required
def ver_cliente(user_id):
    db = get_db()
    cliente_raw = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not cliente_raw:
        flash("Cliente não encontrado.", "danger")
        return redirect(url_for('admin_dashboard'))
        
    cliente = dict(cliente_raw)
    
    if cliente.get('deleted_at'):
        try:
            cliente['deleted_at'] = datetime.strptime(cliente['deleted_at'], '%Y-%m-%d %H:%M:%S.%f')
        except ValueError:
            cliente['deleted_at'] = datetime.strptime(cliente['deleted_at'], '%Y-%m-%d %H:%M:%S')

    historico_raw = db.execute(
        "SELECT id, criado_em, dados_json FROM precificacoes WHERE user_id = ? ORDER BY criado_em DESC", (user_id,)
    ).fetchall()
    
    keys = cliente.keys()
    reset_date = cliente.get('limit_reset_date')
    uso_recente = get_user_usage(db, cliente['id'], reset_date)
    
    historico_list = []
    for row in historico_raw:
        dados = json.loads(row['dados_json'])
        historico_list.append({
            'id': row['id'],
            'criado_em': datetime.strptime(row['criado_em'], '%Y-%m-%d %H:%M:%S'),
            'num_produtos': len(dados)
        })

    return render_template('admin_cliente_detalhes.html', cliente=cliente, historico=historico_list, uso_recente=uso_recente)

@app.route('/admin/cliente/<int:user_id>/update_limit', methods=['POST'])
@login_required
@admin_required
def update_client_limit(user_id):
    try:
        novo_limite = int(request.form.get('novo_limite'))
        if novo_limite < 0:
            flash("O limite não pode ser negativo.", "danger")
        else:
            db = get_db()
            db.execute("UPDATE users SET precificacao_limit = ? WHERE id = ?", (novo_limite, user_id))
            db.commit()
            flash("Limite do cliente atualizado com sucesso!", "success")
    except (ValueError, TypeError):
        flash("Por favor, insira um número válido para o limite.", "danger")
    
    return redirect(url_for('ver_cliente', user_id=user_id))

@app.route('/admin/cliente/<int:user_id>/reset_usage', methods=['POST'])
@login_required
@admin_required
def reset_client_usage(user_id):
    db = get_db()
    now = datetime.now()
    db.execute("UPDATE users SET limit_reset_date = ? WHERE id = ?", (now, user_id))
    db.commit()
    flash("O contador de uso do cliente foi resetado com sucesso!", "success")
    return redirect(url_for('ver_cliente', user_id=user_id))

@app.route('/admin/cliente/<int:user_id>/reactivate', methods=['POST'])
@login_required
@admin_required
def reactivate_client(user_id):
    db = get_db()
    user = db.execute("SELECT nome_completo FROM users WHERE id = ?", (user_id,)).fetchone()

    if not user:
        flash("Cliente não encontrado.", "danger")
        return redirect(url_for('admin_dashboard'))

    db.execute(
        "UPDATE users SET is_deleted = 0, deleted_at = NULL, delete_reason = NULL WHERE id = ?",
        (user_id,)
    )
    db.commit()

    flash(f"A conta de {user['nome_completo']} foi reativada com sucesso!", "success")
    return redirect(url_for('ver_cliente', user_id=user_id))
    
@app.route('/admin/cliente/<int:user_id>/update_status', methods=['POST'])
@login_required
@admin_required
def update_client_status(user_id):
    novo_status = request.form.get('novo_status')
    if novo_status not in ['Safira', 'Esmeralda', 'Diamante']:
        flash("Status inválido selecionado.", "danger")
        return redirect(url_for('ver_cliente', user_id=user_id))

    db = get_db()
    db.execute("UPDATE users SET status_cliente = ? WHERE id = ?", (novo_status, user_id))
    db.commit()
    flash("Status do cliente atualizado com sucesso!", "success")
    return redirect(url_for('ver_cliente', user_id=user_id))

@app.route('/admin/precificacao/<int:prec_id>/resumo')
@login_required
@admin_required
def admin_ver_resumo_precificacao(prec_id):
    db = get_db()
    prec_data = db.execute("""
        SELECT p.id, p.dados_json, p.criado_em, u.nome_completo as user_nome
        FROM precificacoes p JOIN users u ON p.user_id = u.id
        WHERE p.id = ?
    """, (prec_id,)).fetchone()
    if not prec_data:
        flash("Precificação não encontrada.", "danger")
        return redirect(url_for('admin_dashboard'))
    dados = json.loads(prec_data['dados_json'])
    resumo_geral = generate_summary(dados)
    nfe_de_cada_produto = [item.get("Série NF-e", "N/A") for item in dados]
    nfe_counts = Counter(nfe_de_cada_produto)
    nfe_info = sorted(nfe_counts.items())
    criado_em_dt = datetime.strptime(prec_data['criado_em'], '%Y-%m-%d %H:%M:%S')
    return render_template(
        'admin_precificacao_resumo.html',
        prec={
            'id': prec_data['id'],
            'user_nome': prec_data['user_nome'],
            'criado_em': criado_em_dt
        },
        resumo=resumo_geral,
        nfe_info=nfe_info
    )

# --- ROTAS DE DOWNLOAD ---
def mapear_para_shopee(produtos):
    produtos_shopee = []
    for prod in produtos:
        produto_mapeado = {
            '*Nome do Produto': prod.get('Nome'),
            '*Descrição do Produto': prod.get('Nome'),
            '*SKU Principal': prod.get('Código'),
            'Preço': f"{prod.get('Shopee (R$)', 0):.2f}",
            'Estoque': 100,
            '*Código da Categoria': '',
            '*Marca': 'Marca Padrão',
            '*Material': '',
            '*Peso (kg)': '1',
            '*Canal de Envio': ''
        }
        produtos_shopee.append(produto_mapeado)
    return produtos_shopee

@app.route('/baixar/shopee/<int:prec_id>')
@login_required
def baixar_planilha_shopee(prec_id):
    db = get_db()
    query = "SELECT dados_json FROM precificacoes WHERE id = ?"
    params = (prec_id,)
    if not getattr(current_user, 'is_admin', False):
        query += " AND user_id = ?"
        params = (prec_id, current_user.id)
    prec_data = db.execute(query, params).fetchone()
    if prec_data is None:
        flash("Precificação não encontrada.", "danger")
        return redirect(url_for('dashboard'))
    produtos_originais = json.loads(prec_data['dados_json'])
    produtos_mapeados = mapear_para_shopee(produtos_originais)
    df = pd.DataFrame(produtos_mapeados)
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Anuncios Shopee')
    output.seek(0)
    return send_file(
        output,
        download_name="planilha_shopee.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/baixar/completo/<int:prec_id>')
@login_required
def baixar_planilha_completa(prec_id):
    db = get_db()
    query = "SELECT dados_json FROM precificacoes WHERE id = ?"
    params = (prec_id,)
    if not getattr(current_user, 'is_admin', False):
        query += " AND user_id = ?"
        params = (prec_id, current_user.id)
    prec_data = db.execute(query, params).fetchone()
    if prec_data is None:
        flash("Precificação não encontrada ou acesso negado.", "danger")
        return redirect(url_for('dashboard'))
    
    produtos = json.loads(prec_data['dados_json'])
    
    # NOVAS LINHAS ADICIONADAS/MODIFICADAS A PARTIR DAQUI
    # 1. Definir as colunas a serem removidas
    colunas_a_remover = [
        'impostos', 
        'frete_total', 
        'seguro_total', 
        'outros_total', 
        'desc_total',
        'total_nfe_value_from_xml',
        'sum_vProd_from_xml_items'
    ]

    # 2. Processar produtos para remover colunas e formatar valores
    produtos_para_df = []
    for produto in produtos:
        # Cria um novo dicionário sem as colunas a remover
        produto_limpo = {k: v for k, v in produto.items() if k not in colunas_a_remover}
        
        # 3. Formatar valores numéricos para 2 casas decimais
        for key, value in produto_limpo.items():
            # Verifica se o valor é numérico e não é a quantidade ('Qtd') ou identificadores
            if isinstance(value, (int, float)) and key not in ["Qtd", "Série NF-e", "Código", "Nome"]:
                produto_limpo[key] = round(float(value), 2)
            # Também trata strings que podem ser convertidas para float e formata
            elif isinstance(value, str) and value.replace('.', '', 1).isdigit() and key not in ["Série NF-e", "Código", "Nome"]:
                try:
                    produto_limpo[key] = round(float(value), 2)
                except ValueError:
                    pass # Ignora se não for um número válido
        
        produtos_para_df.append(produto_limpo)

    # A linha original 'df = pd.DataFrame(produtos)' é substituída por esta:
    df = pd.DataFrame(produtos_para_df) 
    # FIM DAS NOVAS LINHAS ADICIONADAS/MODIFICADAS
    
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Precificação Completa')
    output.seek(0)
    return send_file(
        output,
        download_name="precificacao_completa.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def mapear_para_ml(produtos):
    produtos_ml = []
    for prod in produtos:
        produto_mapeado = {
            'Título': prod.get('Nome'),
            'Preço': f"{prod.get('Mercado Livre (R$)', 0):.2f}",
            'SKU': prod.get('Código'),
            'Estoque': 100,
            'Tipo de anúncio': 'Clássico',
            'Descrição': prod.get('Nome'),
            'GTIN': ''
        }
        produtos_ml.append(produto_mapeado)
    return produtos_ml

@app.route('/baixar/ml/<int:prec_id>')
@login_required
def baixar_planilha_ml(prec_id):
    db = get_db()
    query = "SELECT dados_json FROM precificacoes WHERE id = ?"
    params = (prec_id,)
    if not getattr(current_user, 'is_admin', False):
        query += " AND user_id = ?"
        params = (prec_id, current_user.id)
    prec_data = db.execute(query, params).fetchone()
    if prec_data is None:
        flash("Precificação não encontrada ou acesso negado.", "danger")
        return redirect(url_for('dashboard'))
    produtos_originais = json.loads(prec_data['dados_json'])
    produtos_mapeados = mapear_para_ml(produtos_originais)
    df = pd.DataFrame(produtos_mapeados)
    output = BytesIO()
    df.to_csv(output, index=False, sep=';', encoding='utf-8-sig')
    output.seek(0)
    return send_file(
        output,
        download_name="planilha_mercado_livre.csv",
        as_attachment=True,
        mimetype="text/csv"
    )

@app.route('/baixar/historico_completo')
@login_required
def baixar_historico_completo():
    db = get_db()
    db_precificacoes = db.execute(
        'SELECT id, criado_em, dados_json FROM precificacoes WHERE user_id = ? ORDER BY criado_em DESC',
        (current_user.id,)
    ).fetchall()
    if not db_precificacoes:
        flash("Você não possui nenhum histórico para baixar.", "info")
        return redirect(url_for('dashboard'))
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for prec in db_precificacoes:
            prec_id = prec['id']
            criado_em = datetime.strptime(prec['criado_em'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
            dados = json.loads(prec['dados_json'])
            df = pd.DataFrame(dados)
            sheet_name = f"ID_{prec_id}_{criado_em}"
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return send_file(
        output,
        download_name="historico_completo_xmarkup.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --- ROTAS DA NOVA FUNCIONALIDADE DE CONCORRENTES ---
# Lógica de scraping básica (manter se houver planos futuros, caso contrário, remover)
def scrape_price_from_url(url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        price_meta = soup.find('meta', itemprop='price')
        if price_meta and price_meta.get('content'):
            try:
                price_str = price_meta['content'].replace('.', '').replace(',', '.').strip()
                return float(re.sub(r'[^\d.]', '', price_str))
            except ValueError:
                pass
        price_tags = soup.find_all(class_=lambda x: x and ('price' in x.lower() or 'valor' in x.lower()))
        for tag in price_tags:
            text = tag.get_text(strip=True)
            match = re.search(r'R\$?\s*(\d{1,3}(?:\.\d{3})*,\d{2})', text)
            if not match:
                match = re.search(r'(\d{1,3}(?:,\d{3})*\.\d{2})', text)
            if match:
                try:
                    price_str = match.group(1).replace('.', '').replace(',', '.')
                    return float(price_str)
                except ValueError:
                    continue
        app.logger.warning(f"Preço não encontrado ou não pôde ser parseado para a URL: {url}")
        return None
    except requests.exceptions.RequestException as e:
        app.logger.error(f"Erro ao aceder a URL {url}: {e}")
        return None
    except Exception as e:
        app.logger.error(f"Erro inesperado no scraping para URL {url}: {e}")
        return None

@app.route('/competitors')
@login_required
def competitor_monitor():
    db = get_db()
    # MODIFICADO: JOIN com a tabela competitors para pegar o nome do concorrente
    competitor_products_raw = db.execute(
        '''SELECT cp.id, cp.product_name, cp.product_url, c.name as competitor_name, cp.marketplace, cp.last_checked_at,
                  (SELECT price FROM competitor_price_history WHERE product_id = cp.id ORDER BY checked_at DESC LIMIT 1) as current_price
           FROM competitor_products cp
           JOIN competitors c ON cp.competitor_profile_id = c.id
           WHERE cp.user_id = ?
           ORDER BY cp.created_at DESC''',
        (current_user.id,)
    ).fetchall()

    competitor_products = []
    for p in competitor_products_raw:
        item = dict(p)
        if item['last_checked_at']:
            item['last_checked_at'] = datetime.strptime(item['last_checked_at'], '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M')
        else:
            item['last_checked_at'] = 'N/A'
        
        if item['current_price'] is None:
            item['current_price'] = 'N/A'
        else:
            item['current_price'] = f"R$ {item['current_price']:.2f}".replace(".", "X").replace(",", ".").replace("X", ",")
            
        competitor_products.append(item)

    return render_template('competitor_monitor.html', products=competitor_products)

# ADICIONADO: Rota para acionar o scraping manualmente para um produto (manter se houver planos futuros)
@app.route('/competitors/scrape/<int:product_id>', methods=['POST'])
@login_required
def scrape_competitor_price(product_id):
    db = get_db()
    product = db.execute(
        'SELECT id, product_url, product_name FROM competitor_products WHERE id = ? AND user_id = ?',
        (product_id, current_user.id)
    ).fetchone()

    if not product:
        flash('Produto de concorrente não encontrado ou acesso não permitido.', 'danger')
        return redirect(url_for('competitor_monitor'))

    scraped_price = scrape_price_from_url(product['product_url'])

    if scraped_price is not None:
        try:
            cursor = db.cursor()
            cursor.execute(
                'INSERT INTO competitor_price_history (product_id, price) VALUES (?, ?)',
                (product_id, scraped_price)
            )
            cursor.execute(
                'UPDATE competitor_products SET last_checked_at = ? WHERE id = ?',
                (datetime.now(), product_id)
            )
            db.commit()
            flash(f"Preço atualizado para '{product['product_name']}' (R$ {scraped_price:.2f}) com sucesso!", "success")
        except Exception as e:
            flash(f"Erro ao guardar o preço raspado: {e}", "danger")
            db.rollback()
    else:
        flash(f"Não foi possível obter o preço para '{product['product_name']}'. Verifique a URL ou a estrutura da página.", "warning")
    
    return redirect(url_for('competitor_product_history', product_id=product_id))

# MODIFICADO: Rota add_competitor_product para usar perfis de concorrentes
@app.route('/competitors/add', methods=['GET', 'POST'])
@login_required
def add_competitor_product():
    db = get_db()
    
    # Pega a lista de perfis de concorrentes do utilizador para o dropdown
    competitor_profiles = db.execute(
        'SELECT id, name FROM competitors WHERE user_id = ? ORDER BY name',
        (current_user.id,)
    ).fetchall()

    if request.method == 'POST':
        product_name = request.form.get('product_name').strip()
        product_url = request.form.get('product_url').strip()
        competitor_profile_id = request.form.get('competitor_profile_id')
        marketplace = request.form.get('marketplace').strip() # 'Site Próprio', 'Mercado Livre', 'Shopee', 'Amazon'
        initial_price = request.form.get('initial_price')

        # Validação básica
        if not product_name or not product_url or not competitor_profile_id or not marketplace:
            flash('Por favor, preencha todos os campos obrigatórios (Nome do Produto, URL, Concorrente e Marketplace).', 'danger')
            return render_template('add_competitor_product.html', competitor_profiles=competitor_profiles)

        try:
            competitor_profile_id = int(competitor_profile_id)
        except ValueError:
            flash('Concorrente selecionado inválido.', 'danger')
            return render_template('add_competitor_product.html', competitor_profiles=competitor_profiles)
            
        cursor = db.cursor()

        try:
            # Verifica se o produto já existe para este concorrente/URL/marketplace
            existing_product = cursor.execute(
                '''SELECT id FROM competitor_products WHERE user_id = ? AND product_url = ? AND marketplace = ?''',
                (current_user.id, product_url, marketplace)
            ).fetchone()

            if existing_product:
                flash(f"Um produto com esta URL e marketplace já está a ser monitorizado para este concorrente.", "warning")
            else:
                # Insere o novo produto de concorrente
                cursor.execute(
                    '''INSERT INTO competitor_products (user_id, competitor_profile_id, product_name, product_url, marketplace)
                       VALUES (?, ?, ?, ?, ?)''',
                    (current_user.id, competitor_profile_id, product_name, product_url, marketplace)
                )
                product_id = cursor.lastrowid

                # Se um preço inicial for fornecido, adiciona-o ao histórico
                if initial_price:
                    price = float(initial_price.replace(',', '.'))
                    cursor.execute(
                        '''INSERT INTO competitor_price_history (product_id, price)
                           VALUES (?, ?)''',
                        (product_id, price)
                    )
                
                db.commit()
                flash('Produto de concorrente adicionado com sucesso!', 'success')
                return redirect(url_for('competitor_monitor'))

        except ValueError:
            flash('Preço inicial inválido. Por favor, insira um número válido.', 'danger')
        except Exception as e:
            flash(f'Ocorreu um erro ao adicionar o produto: {e}', 'danger')
            app.logger.error(f"Erro ao adicionar produto de concorrente: {e}", exc_info=True)
            db.rollback()
        
    return render_template('add_competitor_product.html', competitor_profiles=competitor_profiles)

@app.route('/competitors/delete/<int:product_id>', methods=['POST'])
@login_required
def delete_competitor_product(product_id):
    db = get_db()
    try:
        db.execute('DELETE FROM competitor_products WHERE id = ? AND user_id = ?', (product_id, current_user.id))
        db.commit()
        flash('Produto de concorrente excluído com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao excluir produto: {e}', 'danger')
        db.rollback()
    return redirect(url_for('competitor_monitor'))

@app.route('/competitors/history/<int:product_id>')
@login_required
def competitor_product_history(product_id):
    db = get_db()
    # MODIFICADO: JOIN com a tabela competitors para pegar o nome do concorrente
    product = db.execute('''
        SELECT cp.*, c.name as competitor_name
        FROM competitor_products cp
        JOIN competitors c ON cp.competitor_profile_id = c.id
        WHERE cp.id = ? AND cp.user_id = ?''', (product_id, current_user.id)).fetchone()
        
    if not product:
        flash('Produto não encontrado ou acesso não permitido.', 'danger')
        return redirect(url_for('competitor_monitor'))

    # MODIFICADO AQUI: Lógica mais robusta para parsing da data
    history_raw = db.execute('SELECT price, checked_at FROM competitor_price_history WHERE product_id = ? ORDER BY checked_at ASC', (product_id,)).fetchall() # ASC para o gráfico
    
    history = []
    for h in history_raw:
        checked_at_str = h['checked_at']
        parsed_dt = None
        # Tenta parsear sem microssegundos primeiro (mais comum para CURRENT_TIMESTAMP)
        try:
            parsed_dt = datetime.strptime(checked_at_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            # Se falhar, tenta com microssegundos (menos comum para CURRENT_TIMESTAMP, mas seguro)
            try:
                parsed_dt = datetime.strptime(checked_at_str, '%Y-%m-%d %H:%M:%S.%f')
            except ValueError:
                # Se ambos falharem, loga e continua (ou lida com o erro de outra forma)
                app.logger.error(f"Erro ao parsear data: {checked_at_str}. Formato inesperado.")
                parsed_dt = None # Ou alguma data padrão para evitar quebrar
        
        if parsed_dt:
            history.append({
                'price': h['price'],
                'checked_at': parsed_dt.strftime('%d/%m/%Y %H:%M:%S')
            })
        else:
            history.append({ # Fallback se a data não puder ser parseada
                'price': h['price'],
                'checked_at': checked_at_str # Exibe a string original, mas o gráfico pode falhar
            })
            
    # Prepara dados para gráfico
    chart_labels = [h['checked_at'] for h in history] # Data completa para precisão
    chart_data = [h['price'] for h in history]

    return render_template('competitor_product_history.html', product=product, history=history, chart_labels=chart_labels, chart_data=chart_data)

@app.route('/competitors/compare')
@login_required
def competitor_comparison_graph():
    db = get_db()
    
    # Obter os IDs dos produtos da query string (ex: ?ids=1,2,3)
    product_ids_str = request.args.get('ids', '')
    if not product_ids_str:
        flash('Nenhum produto selecionado para comparação.', 'warning')
        return redirect(url_for('competitor_monitor'))
    
    try:
        product_ids = [int(p_id) for p_id in product_ids_str.split(',') if p_id.isdigit()]
    except ValueError:
        flash('IDs de produto inválidos.', 'danger')
        return redirect(url_for('competitor_monitor'))

    if len(product_ids) < 2:
        flash('Selecione pelo menos dois produtos para comparar.', 'warning')
        return redirect(url_for('competitor_monitor'))

    # Dicionário para armazenar os dados de cada produto
    comparison_data = {}
    
    for p_id in product_ids:
        # Buscar detalhes do produto e garantir que pertence ao utilizador
        product = db.execute('''
            SELECT cp.id, cp.product_name, c.name as competitor_name, cp.marketplace
            FROM competitor_products cp
            JOIN competitors c ON cp.competitor_profile_id = c.id
            WHERE cp.id = ? AND cp.user_id = ?''', (p_id, current_user.id)).fetchone()
            
        if product:
            # Buscar histórico de preços
            history_raw = db.execute(
                'SELECT price, checked_at FROM competitor_price_history WHERE product_id = ? ORDER BY checked_at ASC',
                (p_id,)
            ).fetchall()
            
            history_formatted = []
            for h in history_raw:
                checked_at_str = h['checked_at']
                parsed_dt = None
                try:
                    parsed_dt = datetime.strptime(checked_at_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        parsed_dt = datetime.strptime(checked_at_str, '%Y-%m-%d %H:%M:%S.%f')
                    except ValueError:
                        app.logger.error(f"Erro ao parsear data: {checked_at_str}. Formato inesperado.")
                        parsed_dt = None
                
                if parsed_dt:
                    history_formatted.append({
                        'price': h['price'],
                        'checked_at': parsed_dt.strftime('%Y-%m-%d %H:%M:%S') # Formato padronizado para Chart.js
                    })

            if history_formatted:
                comparison_data[p_id] = {
                    'product_info': dict(product),
                    'history': history_formatted
                }
            else:
                flash(f"Nenhum histórico encontrado para o produto '{product['product_name']}'.", "info")
        else:
            flash(f"Produto com ID {p_id} não encontrado ou não autorizado.", "warning")
            
    if not comparison_data:
        flash('Nenhum produto válido com histórico encontrado para comparação.', 'danger')
        return redirect(url_for('competitor_monitor'))

    return render_template('competitor_comparison_graph.html', comparison_data=comparison_data)

# --- COMANDO CLI PARA MONITORAMENTO DE PREÇOS ---
@app.cli.command("start-price-monitor")
def start_price_monitor():
    """Inicia o monitoramento de preços em background"""
    global price_monitor_scheduler
    
    if price_monitor_scheduler and price_monitor_scheduler.running:
        print("O monitoramento já está em execução!")
        return
    
    db = get_db()
    scraper = SimplePriceScraper(db)
    price_monitor_scheduler = SimpleScrapingScheduler(db, scraper)
    price_monitor_scheduler.start()
    
    print("Monitoramento de preços iniciado! Pressione Ctrl+C para parar.")
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        price_monitor_scheduler.stop()
        print("\nMonitoramento parado.")

# --- ROTAS DE MONITORAMENTO AVANÇADO ---
@app.route('/competitors/dashboard')
@login_required
def competitors_dashboard():
    """Dashboard avançado de monitoramento"""
    db = get_db()
    
    # Estatísticas
    stats = {
        'total_products': db.execute(
            "SELECT COUNT(*) FROM competitor_products WHERE user_id = ?",
            (current_user.id,)
        ).fetchone()[0],
        
        'active_alerts': db.execute(
            """SELECT COUNT(DISTINCT product_id) 
               FROM price_alert_rules 
               WHERE user_id = ? AND enabled = 1""",
            (current_user.id,)
        ).fetchone()[0],
        
        'products_added_this_week': db.execute(
            """SELECT COUNT(*) FROM competitor_products 
               WHERE user_id = ? AND created_at > datetime('now', '-7 days')""",
            (current_user.id,)
        ).fetchone()[0],
        
        'alerts_last_24h': 0,  # Implementar contador de alertas disparados
        'potential_savings': 0,  # Calcular baseado em diferenças de preço
        'prediction_accuracy': 85  # Placeholder - implementar cálculo real
    }
    
    # Produtos com dados agregados
    products_query = """
        SELECT 
            cp.*,
            c.name as competitor_name,
            (SELECT price FROM competitor_price_history 
             WHERE product_id = cp.id 
             ORDER BY checked_at DESC LIMIT 1) as current_price,
            (SELECT price FROM competitor_price_history 
             WHERE product_id = cp.id 
             AND checked_at > datetime('now', '-1 day')
             ORDER BY checked_at ASC LIMIT 1) as price_24h_ago
        FROM competitor_products cp
        JOIN competitors c ON cp.competitor_profile_id = c.id
        WHERE cp.user_id = ?
        ORDER BY cp.last_checked_at DESC
    """
    
    products = []
    for row in db.execute(products_query, (current_user.id,)).fetchall():
        product = dict(row)
        
        # Calcula variação 24h
        if product['price_24h_ago'] and product['current_price']:
            product['price_change_24h'] = (
                (product['current_price'] - product['price_24h_ago']) 
                / product['price_24h_ago'] * 100
            )
        else:
            product['price_change_24h'] = None
            
        # Adiciona mais dados
        product['trend'] = 'stable'  # Placeholder
        product['predicted_price'] = None  # Placeholder
        product['my_price'] = None  # Buscar do sistema de precificação
        product['alert_active'] = False  # Verificar alertas ativos
        product['out_of_stock'] = False  # Verificar estoque
        
        products.append(product)
    
    # Lista de concorrentes
    competitors = db.execute(
        "SELECT * FROM competitors WHERE user_id = ? ORDER BY name",
        (current_user.id,)
    ).fetchall()
    
    return render_template(
        'competitor_monitor_advanced.html',
        products=products,
        stats=stats,
        competitors=competitors
    )

@app.route('/test-scraping/<int:product_id>')
@login_required
def test_scraping(product_id):
    """Testa o scraping de um produto específico"""
    db = get_db()
    
    # Verifica se o produto pertence ao usuário
    product = db.execute(
        """SELECT cp.*, c.name as competitor_name
           FROM competitor_products cp
           JOIN competitors c ON cp.competitor_profile_id = c.id
           WHERE cp.id = ? AND cp.user_id = ?""",
        (product_id, current_user.id)
    ).fetchone()
    
    if not product:
        flash('Produto não encontrado.', 'danger')
        return redirect(url_for('competitor_monitor'))
    
    # Faz o scraping
    scraper = SimplePriceScraper(db)
    result = scraper.scrape_product(product_id, product['product_url'], product['marketplace'])
    scraper.close()
    
    if result:
        flash(f"Scraping bem-sucedido! Preço: R$ {result['price']:.2f}", 'success')
        
        # Salva o resultado
        db.execute(
            "INSERT INTO competitor_price_history (product_id, price, additional_data) VALUES (?, ?, ?)",
            (product_id, result['price'], json.dumps(result))
        )
        db.execute(
            "UPDATE competitor_products SET last_checked_at = ? WHERE id = ?",
            (datetime.now(), product_id)
        )
        db.commit()
    else:
        flash('Não foi possível obter o preço. Verifique a URL.', 'warning')
    
    return redirect(url_for('competitor_product_history', product_id=product_id))

@app.route('/api/competitor/product/<int:product_id>/analytics')
@login_required
def api_competitor_analytics(product_id):
    """API para analytics básico de produto"""
    db = get_db()
    
    # Verifica permissão
    product = db.execute(
        "SELECT * FROM competitor_products WHERE id = ? AND user_id = ?",
        (product_id, current_user.id)
    ).fetchone()
    
    if not product:
        return jsonify({"error": "Produto não encontrado"}), 404
    
    # Busca histórico de preços
    history = db.execute(
        """SELECT price, checked_at 
           FROM competitor_price_history 
           WHERE product_id = ? 
           ORDER BY checked_at DESC 
           LIMIT 30""",
        (product_id,)
    ).fetchall()
    
    if not history:
        return jsonify({"error": "Sem dados históricos"}), 404
    
    prices = [h['price'] for h in history]
    dates = [h['checked_at'] for h in history]
    
    # Calcula estatísticas básicas
    analysis = {
        "product_id": product_id,
        "current_price": prices[0] if prices else 0,
        "stats": {
            "mean": sum(prices) / len(prices) if prices else 0,
            "min": min(prices) if prices else 0,
            "max": max(prices) if prices else 0,
            "count": len(prices)
        },
        "history": [
            {"price": p, "date": d} for p, d in zip(prices, dates)
        ]
    }
    
    return jsonify(analysis)

@app.route('/api/competitor/alerts/configure', methods=['POST'])
@login_required
def api_configure_alerts():
    """Configura alertas para um produto"""
    data = request.get_json()
    db = get_db()
    
    # Validação
    product_id = data.get('product_id')
    if not product_id:
        return jsonify({"error": "Product ID required"}), 400
    
    # Verifica se o produto pertence ao usuário
    product = db.execute(
        "SELECT * FROM competitor_products WHERE id = ? AND user_id = ?",
        (product_id, current_user.id)
    ).fetchone()
    
    if not product:
        return jsonify({"error": "Produto não encontrado"}), 404
    
    # Salva configurações
    alerts = data.get('alerts', {})
    
    # Remove alertas existentes
    db.execute(
        "DELETE FROM price_alert_rules WHERE user_id = ? AND product_id = ?",
        (current_user.id, product_id)
    )
    
    # Cria novos alertas
    for alert_type, config in alerts.items():
        if isinstance(config, dict) and config.get('enabled'):
            db.execute(
                """INSERT INTO price_alert_rules 
                   (user_id, product_id, alert_type, threshold, enabled) 
                   VALUES (?, ?, ?, ?, ?)""",
                (current_user.id, product_id, alert_type, 
                 config.get('threshold', 5.0), True)
            )
        elif config == True:  # Para alertas simples como out_of_stock
            db.execute(
                """INSERT INTO price_alert_rules 
                   (user_id, product_id, alert_type, enabled) 
                   VALUES (?, ?, ?, ?)""",
                (current_user.id, product_id, alert_type, True)
            )
    
    # Atualiza frequência
    frequency = data.get('frequency', 720)
    db.execute(
        "UPDATE competitor_products SET monitoring_frequency = ? WHERE id = ?",
        (frequency, product_id)
    )
    
    db.commit()
    
    return jsonify({"status": "success"})


# --- COMANDO CLI PARA PROMOVER UTILIZADOR A ADMIN ---
@app.cli.command("promote-user")
@click.argument("email")
def promote_user_command(email):
    """Atribui privilégios de administrador a um utilizador existente."""
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    
    if user is None:
        print(f"Erro: Utilizador com o e-mail '{email}' não encontrado.")
        return

    if user['is_admin']:
        print(f"Aviso: O utilizador '{email}' já é um administrador.")
        return

    db.execute("UPDATE users SET is_admin = 1 WHERE email = ?", (email,))
    db.commit()
    print(f"Sucesso: O utilizador '{email}' foi promovido a administrador.")
    
if __name__ == "__main__":
    app.run(debug=True)